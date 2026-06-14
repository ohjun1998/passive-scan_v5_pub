#!/usr/bin/env python3
import os
import glob
import re
import json
import subprocess
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def escape_formula(value):
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
        return "'" + value
    return value

# 💡 상대 경로를 'http://도메인/경로' 로 조립해주는 함수
def make_absolute(url, domain):
    if url.startswith('http://') or url.startswith('https://'):
        return url
    elif url.startswith('//'):
        return f"https:{url}"
    elif url.startswith('/'):
        return f"https://{domain}{url}"
    else:
        return f"https://{domain}/{url}"

def get_status_color(status):
    status_str = str(status)
    if status_str.startswith('2'): return '28A745' # Green (성공)
    if status_str.startswith('3'): return '17A2B8' # Blue (리다이렉트)
    if status_str.startswith('4'): return 'FD7E14' # Orange (권한없음, Not Found)
    if status_str.startswith('5'): return 'DC3545' # Red (서버 에러)
    return '6C757D' # Gray (Dead / 알수없음)

def build_advanced_excel_report():
    print("[+] Initializing Modern Premium Excel Dashboard Engine...", flush=True)
    if not os.path.exists('targets.txt'): return
    
    with open('targets.txt', 'r') as f:
        targets = [line.strip() for line in f if line.strip()]

    js_url_converter = {}
    mapping_files = glob.glob('results/*_js_mapping.txt')
    for mf in mapping_files:
        try:
            with open(mf, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    if '\t' in line:
                        safe_fname, original_url = line.strip().split('\t', 1)
                        js_url_converter[safe_fname] = original_url
        except: pass

    matrix_data = {domain: {} for domain in targets}
    txt_files = glob.glob('results/**/*.*', recursive=True) + glob.glob('results/*.*')
    txt_files = [f for f in txt_files if os.path.isfile(f)]

    # 1. 파일 이름에서 명확한 도메인 식별 후 파싱 진행
    for file_path in txt_files:
        filename = os.path.basename(file_path).lower()
        
        match = re.match(r'^(.*)_(linkfinder|trufflehog|gau|waybackurls)\.txt$', filename)
        if not match: continue
        
        current_domain = match.group(1)
        if current_domain not in targets: continue

        if 'linkfinder' in filename or 'jsluice' in filename: source_tool = 'LinkFinder'
        elif 'trufflehog' in filename: source_tool = 'TruffleHog'
        elif 'waybackurls' in filename: source_tool = 'Waybackurls'
        elif 'gau' in filename: source_tool = 'GAU'
        else: continue

        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    line_str = line.strip()
                    if not line_str or line_str.startswith('#'): continue
                    line_str = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', line_str)
                    if not line_str: continue
                    
                    js_file = "Passive Archive"
                    if '\t' in line_str:
                        parts = line_str.split('\t', 1)
                        js_file, raw_url = parts[0], parts[1]
                    else:
                        raw_url = line_str
                    
                    if js_file in js_url_converter:
                        js_file = js_url_converter[js_file]
                    
                    # 💡 절대 경로 조립 & 서브도메인 엄격한 차단 로직
                    abs_url = make_absolute(raw_url, current_domain)
                    parsed_netloc = urlparse(abs_url).netloc.split(':')[0]
                    if parsed_netloc != current_domain:
                        continue # banana.test.com 같은 서브도메인은 여기서 완벽히 버려짐

                    if abs_url not in matrix_data[current_domain]:
                        matrix_data[current_domain][abs_url] = {"tools": set(), "files": set()}
                    matrix_data[current_domain][abs_url]["tools"].add(source_tool)
                    if source_tool in ['LinkFinder', 'TruffleHog']:
                        matrix_data[current_domain][abs_url]["files"].add(js_file)
        except: pass

    # 2. 💡 모든 추출된 절대 경로를 대상으로 httpx 스텔스 라이브 검증
    all_urls = set()
    for domain, url_map in matrix_data.items():
        all_urls.update(url_map.keys())

    status_codes = {}
    if all_urls:
        print(f"[+] Launching Stealth httpx on {len(all_urls)} endpoints...", flush=True)
        with open('httpx_targets.txt', 'w') as f:
            for u in all_urls: f.write(u + '\n')
            
        # 스텔스 규칙: 무작위 브라우저, 5요청/초 제한, 4초 연결해제
        os.system("httpx -l httpx_targets.txt -sc -random-agent -rl 5 -timeout 4 -retries 1 -json -o httpx_results.json -silent")
        
        if os.path.exists('httpx_results.json'):
            with open('httpx_results.json', 'r') as f:
                for line in f:
                    try:
                        data = json.loads(line.strip())
                        status_codes[data.get('url')] = data.get('status_code', 'Dead')
                    except: pass

    # 3. 엑셀 워크북 생성 및 스타일 정의
    wb = Workbook()
    font_header = Font(name='Malgun Gothic', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='2F3542', end_color='2F3542', fill_type='solid') 
    fill_zebra = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')  
    font_data = Font(name='Malgun Gothic', size=10, color='333333')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    thin_side = Side(border_style="thin", color="E0E0E0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws_dash = wb.active
    ws_dash.title = "Summary Dashboard"
    dash_headers = ["No", "Target Domain", "Total URLs", "jsluice 추출 개수", "TruffleHog 탐지 개수"]
    ws_dash.append(dash_headers)
    for c, _ in enumerate(dash_headers, 1):
        cell = ws_dash.cell(row=1, column=c)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_center; cell.border = thin_border

    ws_high = wb.create_sheet(title="High Risk Targets")
    high_headers = ["No", "Source Tool", "Found in JS File", "Status", "Domain", "High Risk URL / Endpoint", "Risk Reason"] 
    ws_high.append(high_headers)
    for c, _ in enumerate(high_headers, 1):
        cell = ws_high.cell(row=1, column=c)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_center; cell.border = thin_border

    high_risk_keywords = ['config', '.env', 'xml', 'json', 'secret', 'api/v', 'token', 'admin', 'password', 'key', 'credential', 'mysql']
    dash_idx = high_risk_idx = 2  

    for domain, url_map in matrix_data.items():
        sheet_title = re.sub(r'[\\/\?\*\:\[\]]', '_', domain)[:30]

        passive_count = sum(1 for data in url_map.values() if 'Waybackurls' in data["tools"] or 'GAU' in data["tools"])
        jsluice_count = sum(1 for data in url_map.values() if 'LinkFinder' in data["tools"])
        trufflehog_count = sum(1 for data in url_map.values() if 'TruffleHog' in data["tools"])
        
        ws_dash.append([dash_idx - 1, escape_formula(domain), passive_count, jsluice_count, trufflehog_count])
        for col_num in range(1, 6):
            cell = ws_dash.cell(row=dash_idx, column=col_num)
            cell.font = font_data; cell.border = thin_border
            if col_num in [1, 2]: cell.alignment = align_center if col_num == 1 else align_left
            if col_num == 2 and url_map:
                cell.hyperlink = f"#'{sheet_title}'!A1"
                cell.font = Font(name='Malgun Gothic', size=10, color='0056B3', underline='single')
        dash_idx += 1

        if not url_map: continue

        ws = wb.create_sheet(title=sheet_title)
        ws.append(["No", "Source Tool", "Found in JS File", "Status", "Target Absolute URL"]) 
        for c in range(1, 6):
            cell = ws.cell(row=1, column=c)
            cell.font = font_header; cell.fill = fill_header; cell.alignment = align_center; cell.border = thin_border

        for sub_idx, (url, data) in enumerate(sorted(url_map.items()), 1):
            if sub_idx > 1048500: break
            tools_str = ", ".join(sorted(list(data["tools"])))
            files_str = ", ".join(sorted(list(data["files"]))) if data["files"] else "-"
            
            # 💡 Httpx 매핑 데이터 연동
            current_status = status_codes.get(url, 'Dead')
            
            ws.append([sub_idx, escape_formula(tools_str), escape_formula(files_str), current_status, escape_formula(url)]) 
            row_num = sub_idx + 1
            
            for c in range(1, 6):
                cell = ws.cell(row=row_num, column=c)
                cell.font = font_data; cell.border = thin_border
                if (row_num % 2) == 1: cell.fill = fill_zebra
                
                # 상태 코드에 따른 컬러 입히기
                if c == 4:
                    color_hex = get_status_color(current_status)
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                    cell.font = Font(name='Malgun Gothic', size=10, bold=True, color='FFFFFF')
                    cell.alignment = align_center
                elif c in [3, 5]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center

            is_high_risk = False
            reason = ""
            if 'TruffleHog' in data["tools"]:
                is_high_risk, reason = True, "TruffleHog 검증 완료 민감 키(Secret) 유출 징후"
            else:
                matched = [key for key in high_risk_keywords if key in url.lower()]
                if matched:
                    is_high_risk, reason = True, f"민감 키워드 감지 ({', '.join(matched)})"
                    
            if is_high_risk:
                ws_high.append([high_risk_idx - 1, escape_formula(tools_str), escape_formula(files_str), current_status, escape_formula(domain), escape_formula(url), escape_formula(reason)]) 
                for c in range(1, 8):
                    cell = ws_high.cell(row=high_risk_idx, column=c)
                    cell.font = font_data; cell.border = thin_border
                    if (high_risk_idx % 2) == 1: cell.fill = fill_zebra
                    
                    if c == 4:
                        color_hex = get_status_color(current_status)
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                        cell.font = Font(name='Malgun Gothic', size=10, bold=True, color='FFFFFF')
                        cell.alignment = align_center
                    elif c in [3, 5, 6, 7]: cell.alignment = align_left
                    else: cell.alignment = align_center
                high_risk_idx += 1

    for sheet in wb.worksheets:
        for col_idx, col in enumerate(sheet.columns, 1):
            col_letter = get_column_letter(col_idx)
            header_value = sheet.cell(row=1, column=col_idx).value
            
            if header_value in ["Target Absolute URL", "High Risk URL / Endpoint"]:
                sheet.column_dimensions[col_letter].width = 80  
            elif header_value == "Found in JS File":
                sheet.column_dimensions[col_letter].width = 50  
            elif header_value == "Risk Reason":
                sheet.column_dimensions[col_letter].width = 40  
            elif header_value == "Status":
                sheet.column_dimensions[col_letter].width = 12
            else:
                sheet.column_dimensions[col_letter].width = 20

    ws_dash.column_dimensions['B'].width = 35

    os.makedirs('reports', exist_ok=True)
    wb.save('reports/passive_recon_report_v1.xlsx')

if __name__ == '__main__':
    build_advanced_excel_report()
